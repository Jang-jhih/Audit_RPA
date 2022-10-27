import pandas as pd
import os
from openpyxl.utils.dataframe import dataframe_to_rows
import pandas as pd
import openpyxl
import os
import datetime

import warnings
warnings.filterwarnings("ignore")
#%% 
# 用途：將檔案整理成底稿主表
# 傳入參數：path；路徑
# 回傳：df

def MadeWP(All_AuditItem,path):
    All_File_TXT = All_AuditItem[All_AuditItem['files'].str.contains('txt')]
     
    
    #%%提取TXT內容
    All_content = CreatTxtContent(txtlist=All_File_TXT['原始路徑'].tolist())
    
    #%% WorkPaper SHEET
        
    All_content=pd.DataFrame({'原始路徑':All_File_TXT['原始路徑'].tolist(),
                  '文件內容':All_content})
    
    All_AuditItem = pd.merge(All_AuditItem,All_content,on = '原始路徑' ,how ='outer')
    
    
    #%% PNG SHEET
    
    All_png = All_AuditItem[All_AuditItem['files'].str.contains('png', na=False)]
    
    #排序圖片
    All_png = All_png.sort_values('files')
    
    
    All_png =ProcessCell(All_png)
    All_AuditItem =ProcessCell(All_AuditItem)
    
    
    HYPERLINK_png = CreatHyperlink(FilesList = All_png['files'].tolist(),RootList = All_png['root'].tolist())
    
    All_png['HYPERLINK'] = HYPERLINK_png
    
    HYPERLINK_workpaper = CreatHyperlink(FilesList = All_AuditItem['files'].tolist(),RootList = All_AuditItem['root'].tolist())
    All_AuditItem['HYPERLINK'] = HYPERLINK_workpaper
    
    
    workbook =openpyxl.load_workbook(filename='VBA.xlsm', read_only=False, keep_vba=True)
    
    #建立sheet
    worksheet = workbook.create_sheet("附件彙總",1)
    
    
    wspng = workbook.create_sheet("PNG檔彙總",2)
    
    #塞圖片
    inputPNG(sheet=wspng,PngPathList=All_png['原始路徑'].tolist())
    
    
    All_png = All_png[['主項目', '子項目1', '子項目2',  'HYPERLINK']]
    All_png.columns = ['主項目', '子項目1', '子項目2',  '檔案名稱' ]
    
    All_AuditItem = All_AuditItem[['主項目', '子項目1', '子項目2',  'HYPERLINK','ReviseTime','文件內容']]
    All_AuditItem.columns = ['主項目', '子項目1', '子項目2',  '檔案名稱','ReviseTime','文件內容']
    
    All_AuditItem['查核人員意見(x : 無)'] = ''
    All_AuditItem['備註'] = ''
 
 
 
 
    #%%
    #塞資料到sheet
    All_AuditItem = All_AuditItem.fillna('')
 
 
    All_AuditItem.columns
    
    for Serise in ['主項目', '子項目1', '子項目2',  '文件內容', '查核人員意見(x : 無)',
           '備註']:
        All_AuditItem[Serise] = clean(All_AuditItem[Serise])
 
 
 
    inputDF(worksheet=wspng,df=All_png,header=True,index = False)
    inputDF(worksheet=worksheet,df=All_AuditItem)
 
 
#%%


    #微調png sheet
    wspng.insert_cols(1,1)
    wspng.cell(row=1, column=1).value = '圖片'

    #%% 調整樣式
    for worksheet,wspng in zip(worksheet['D'],wspng['E']):
        worksheet.style = "Hyperlink"
        wspng.style = "Hyperlink"
    
    
    
    #%%存檔
    
    from datetime import datetime
    
    today = datetime.now()
    today = datetime.strftime(today, "%Y%m%d")
    
    
    
    department = path.split('\\')[-1].split('_')[-1]
    workbook.save(f'WorkBook_{department}_{today}.xlsm')
    # workbook.save(f'WorkBook.xlsm')



def PrintFilePath (path):

    
    allroot =[]
    allfiles=[]
    
    for allpath in [os.path.join(path,_) for _ in os.listdir(path)]:
        for root, dirs, files in os.walk(allpath):
            files.sort(reverse=True)
            allroot.append(root)
            allfiles.append(files)
            
    df = pd.DataFrame({'root':allroot,
                       'files':allfiles})
    df = df.explode('files')
    df['原始路徑'] = df['root'] +'\\' +df['files']
    
    df = df[df['files'].notnull()]
    
    subitem= [_.replace(f"{path}\\",'') for _ in df['原始路徑'].tolist()]
    
    AuditItem = [_.replace('，',',').split('\\') for _ in subitem]
    
    AuditItem = pd.DataFrame(AuditItem)
    
    AuditItem.columns = [f'子項目{_}' if _ > 0 else '主項目' for _ in AuditItem.columns.tolist() ]
    AuditItem.reset_index(drop = True ,inplace =True)
    df.reset_index(drop = True ,inplace =True)
    df = pd.concat([AuditItem,df] ,axis=1)
    
    #取得修改時間
    ReviseTime = []
    for file in df['原始路徑'].tolist():
        unix_time=os.path.getmtime(file)
        datetimeObj = datetime.datetime.fromtimestamp(unix_time)
        ReviseTime.append(datetimeObj)
    
    df['ReviseTime'] = ReviseTime
    
    return df

#%% 
# 用途：把df塞入sheet
# 傳入參數：worksheet=被塞的sheet,header=是否有表頭,index=是否有index
# 回傳：


def inputDF(worksheet,df,header=True,index=False):
    for row in dataframe_to_rows(df = df, index=index, header=header):
        worksheet.append(row)
        
#%% 
# 用途：清除有副檔名的cell
# 傳入參數：df = df
# 回傳：清除後的df
def ProcessCell(df):
    for filmename in ['.png','.txt','.db','.pdf','xlsm','xlxm','.xlsx','jpg']:
        for column in df.columns[:3]:
        
            filterbool = df[column].str.contains(filmename, na=False)
            
            df.loc[(filterbool), column] = ''
    
    df.reset_index(drop = True ,inplace =True)
    
    return df

#%%
def clean(Serise):
    Serise = Serise.str.replace(',','，').str.replace(' ','').str.replace('\n','').str.replace('，','，\n')
    return Serise
#%%
def inputPNG(sheet,PngPathList):

    for i,png in zip(range(2,len(PngPathList)),PngPathList):
        #塞圖片
        # PNG's sheet columns
        pngplace = 'A'
        
        # PNG Multiple
        Multiple = 2
        
        img= openpyxl.drawing.image.Image(png)
        resize = 0.35278*Multiple
        height =720*resize
        width =1280*resize/2
        img.width, img.height = (width,height)
        sheet.add_image(img,f'{pngplace}{i}')
    
        sheet.row_dimensions[i].height = height*0.8
        sheet.column_dimensions[pngplace].width = width/7
        
        
#%%建立檔案超連結
def CreatHyperlink(FilesList,RootList):
    HYPERLINK = []
    for cell,link in zip(FilesList,RootList):
        HYPERLINK.append(f'==HYPERLINK("{link}", "{cell}")')
    return HYPERLINK

#%%提取TXT內容
def CreatTxtContent(txtlist):
    All_content = []
    
    for TXTcontent in txtlist:
        content = open(TXTcontent,'r', encoding='UTF-8').read()
        All_content.append(content)
    
    try:
        All_content = [_.split('檔案說明')[1] for _ in All_content]
    except:
        print('沒有檔案說明.txt')
        
    return All_content